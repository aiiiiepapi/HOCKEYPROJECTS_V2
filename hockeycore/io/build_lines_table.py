"""Friendly 30-second lines workbook (Seb 2026-07-25).

One filterable table: time (30s steps, 15:00->3:00) x state x coach tier,
probabilities as %, fair (0% EV) lines + 10%-EV American lines, color-coded,
frozen header, autofilter. Pure values (no formulas) — cheap-session rebuildable:
    python3 hockeycore/io/build_lines_table.py
"""
import json
from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

ROOT = Path(__file__).resolve().parents[2]
DER = ROOT / "data" / "derived"
OUT = ROOT / "live_model" / "lines_table_30s.xlsx"
EDGE = 0.10

STATE_LABEL = {
    "not_pulled_EV": "Net in — 5v5",
    "not_pulled_PP": "Net in — POWER PLAY",
    "pulled": "GOALIE PULLED",
}
STATE_FILL = {
    "not_pulled_EV": PatternFill("solid", fgColor="F5F5F5"),
    "not_pulled_PP": PatternFill("solid", fgColor="DCE9F7"),
    "pulled": PatternFill("solid", fgColor="FCE4D6"),
}
TIER_LABEL = {0.55: "0.55 very passive", 0.80: "0.80 passive", 1.00: "1.00 league avg",
              1.30: "1.30 aggressive", 1.85: "1.85 very aggressive", 2.40: "2.40 extreme"}
MARKETS = [("P_total_ge1", "Game total OVER"), ("P_leader_ge1", "Leader TT OVER"),
           ("P_margin_ge4", "Leader -3.5")]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF")
ARIAL = Font(name="Arial")
BOLD = Font(name="Arial", bold=True)
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))


def american(p, edge=EDGE):
    b = (1 + edge - p) / p
    return f"+{b*100:.0f}" if b >= 1 else f"-{100/b:.0f}"


def main():
    grid = json.load(open(DER / "pricing_grid_dense.json"))
    profiles = json.load(open(DER / "coach_profiles.json"))
    Rs = np.array(sorted({r["R"] for r in grid}))
    by = {(r["state"], r["tier"], r["R"]): r for r in grid}
    states = ["not_pulled_EV", "not_pulled_PP"]  # pulled rows removed (Seb: useless, no lines offered)
    tiers = sorted({r["tier"] for r in grid})
    times = list(range(900, 179, -30))

    wb = Workbook()
    ws = wb.active
    ws.title = "LINES"
    ws.sheet_view.showGridLines = False

    ws["A1"] = ("3-GOAL GAP, 3rd PERIOD — model probability + fair line (0% EV) + the line that pays +10% EV "
                "(bet only at that number or BETTER). Filter any column. NO-GO for real money; "
                "leader market = Overs only.")
    ws["A1"].font = Font(name="Arial", bold=True, size=11)
    ws.merge_cells("A1:L1")

    headers = ["Time left", "Situation", "Coach type"]
    for _, lbl in MARKETS:
        headers += [f"{lbl} — prob", f"{lbl} — TRUE line", f"{lbl} — line @10%"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    row = 3
    for st in states:
        for tr in tiers:
            curves = {}
            for mk, _ in MARKETS:
                ys = np.array([by[(st, tr, R)][mk] for R in Rs])
                curves[mk] = {t: float(np.interp(t, Rs, ys)) for t in times}
            for t in times:
                vals = [f"{t//60}:{t%60:02d}", STATE_LABEL[st], TIER_LABEL[tr]]
                cells = []
                for mk, _ in MARKETS:
                    p = curves[mk][t]
                    cells += [p, american(p, edge=0.0), american(p)]
                for j, v in enumerate(vals + cells, 1):
                    c = ws.cell(row=row, column=j, value=v)
                    c.font = ARIAL
                    c.border = THIN
                    if j <= 3:
                        c.fill = STATE_FILL[st]
                    if j in (4, 7, 10):
                        c.number_format = "0.0%"
                    if j in (5, 6, 8, 9, 11, 12):
                        c.font = BOLD
                        c.alignment = Alignment(horizontal="center")
                row += 1

    last = row - 1
    ws.auto_filter.ref = f"A2:L{last}"
    ws.freeze_panes = "A3"
    for col, w in zip("ABCDEFGHIJKL", (9, 22, 20, 14, 13, 13, 14, 13, 13, 14, 13, 13)):
        ws.column_dimensions[col].width = w
    # green->red scale on each probability column (high prob = green)
    for col in ("D", "G", "J"):
        ws.conditional_formatting.add(
            f"{col}3:{col}{last}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))


    # ---------- COACH GUIDE ----------
    cg = wb.create_sheet("COACH GUIDE", 1)
    cg.sheet_view.showGridLines = False
    cg["A1"] = ("WHO PULLS, AND WHEN — EXPECTED pull % = model estimate on a clean 5v5 chance: "
                "recency-weighted record + league prior (3/3 prices ABOVE league, cold streaks count). "
                "PP pulls excluded (not bettable). Last 3 = newest chances, P=pulled n=no. "
                "TEAM LINES tab prices each team at its own coach's % and timing.")
    cg["A1"].font = Font(name="Arial", bold=True, size=11)
    cg.merge_cells("A1:H1")
    ghdr = ["Team", "Coach", "EXPECTED pull % (model)", "Record", "Last 3", "Range (95%)",
            "Typical pull time left", "Notes"]
    for j, h in enumerate(ghdr, 1):
        c = cg.cell(row=2, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    def tier_of(m):
        best = min(TIER_LABEL, key=lambda t: abs(t - m))
        return TIER_LABEL[best]

    # current coach = latest last_seen per team
    cur = {}
    for p in profiles:
        if p["team"] not in cur or p["last_seen"] > cur[p["team"]]["last_seen"]:
            cur[p["team"]] = p
    rowg = 3
    ordered = sorted(profiles, key=lambda p: (p["team"], p is not cur.get(p["team"]), p["coach"]))
    for p in ordered:
        is_cur = cur.get(p["team"]) is p
        pp_only = any("ONLY on PP" in f for f in p["flags"])
        notes = list(p["flags"])
        med = p["median_pull_R"]
        timing = f"{med//60}:{med%60:02d}" if med and p["n_pulls"] >= 3 else "~4:19 (league, few pulls)"
        vals = [p["team"], p["coach"] + ("" if is_cur else " (former)"),
                p["expected_pull_pct"], f'{p["clear_taken"]}/{p["clear_chances"]}',
                p.get("last3", "-"),
                (f'{p["band"][0]:.0%}-{p["band"][1]:.0%}' if p.get("band") else "n/a"),
                timing, "; ".join(notes)]
        for j, v in enumerate(vals, 1):
            c = cg.cell(row=rowg, column=j, value=v)
            c.font = BOLD if (j == 2 and is_cur) else ARIAL
            c.border = THIN
            if j == 3:
                c.number_format = "0%"
            if not is_cur:
                c.font = Font(name="Arial", color="999999")
        rowg += 1
    lastg = rowg - 1
    cg.auto_filter.ref = f"A2:H{lastg}"
    cg.freeze_panes = "A3"
    for col, w in zip("ABCDEFGH", (7, 26, 12, 9, 12, 15, 30, 46)):
        cg.column_dimensions[col].width = w
    cg.conditional_formatting.add(
        f"C3:C{lastg}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.64, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B"))


    # ---------- TEAM LINES (per current coach, manual grid) ----------
    import bisect as _bi
    mg = json.load(open(DER / "pricing_grid_manual.json"))
    PGRID = sorted({r["pullpct"] for r in mg}); SGRID = sorted({r["shift"] for r in mg})
    RGRID = sorted({r["R"] for r in mg})
    MGET = {(r["R"], r["pullpct"], r["shift"]): r for r in mg}
    def _tri(R, pp, sh, mk):
        def brk(g, x):
            x = min(max(x, g[0]), g[-1])
            i = min(_bi.bisect_right(g, x), len(g)-1)
            lo = g[max(i-1,0)]; hi = g[i] if g[i] != lo else g[min(i, len(g)-1)]
            if hi == lo: return lo, hi, 0.0
            return lo, hi, (x-lo)/(hi-lo)
        r0,r1,wr = brk(RGRID,R); p0,p1,wp = brk(PGRID,pp); s0,s1,ws = brk(SGRID,sh)
        v = 0.0
        for ri,wi in ((r0,1-wr),(r1,wr)):
            for pi,wj in ((p0,1-wp),(p1,wp)):
                for si,wk in ((s0,1-ws),(s1,ws)):
                    if wi*wj*wk: v += wi*wj*wk*MGET[(ri,pi,si)][mk]
        return v
    tl = wb.create_sheet("TEAM LINES", 1)
    tl.sheet_view.showGridLines = False
    tl["A1"] = ("EACH TEAM AT ITS OWN COACH'S NUMBERS (5v5, net still in) — expected pull % + his timing, "
                "priced through the model. Filter your team. Lines = worst price that still pays +10% EV.")
    tl["A1"].font = Font(name="Arial", bold=True, size=11)
    tl.merge_cells("A1:J1")
    thdr = ["Team", "Coach", "Pull% used", "Time left"]
    for _, lbl in MARKETS: thdr += [f"{lbl} — prob", f"{lbl} — line @10%"]
    for j, h in enumerate(thdr, 1):
        c = tl.cell(row=2, column=j, value=h); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    tr = 3
    for team in sorted(cur):
        p = cur[team]
        pp_used = p["expected_pull_pct"]
        med = p["median_pull_R"]
        sh = (med - 259) if (med and p["n_pulls"] >= 3) else 0
        sh = max(-60, min(180, sh))
        for t in range(900, 179, -30):
            vals = [team, p["coach"], pp_used, f"{t//60}:{t%60:02d}"]
            for mk, _ in MARKETS:
                pv = _tri(t, pp_used, sh, mk)
                vals += [pv, american(pv)]
            for j, v in enumerate(vals, 1):
                c = tl.cell(row=tr, column=j, value=v)
                c.font = ARIAL; c.border = THIN
                if j == 3: c.number_format = "0%"
                if j in (5,7,9): c.number_format = "0.0%"
                if j in (6,8,10): c.font = BOLD; c.alignment = Alignment(horizontal="center")
            tr += 1
    lastt = tr - 1
    tl.auto_filter.ref = f"A2:J{lastt}"
    tl.freeze_panes = "A3"
    for col, w in zip("ABCDEFGHIJ", (7, 24, 10, 9, 13, 12, 13, 12, 13, 12)):
        tl.column_dimensions[col].width = w
    for col in ("E","G","I"):
        tl.conditional_formatting.add(f"{col}3:{col}{lastt}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))


    # ---------- NO-PULL LOG (every non-pull, classified) ----------
    cwr = json.load(open(DER / "clean_window_instances.json"))
    npl = wb.create_sheet("NO-PULL LOG")
    npl.sheet_view.showGridLines = False
    npl["A1"] = ("EVERY NO-PULL INSTANCE, CLASSIFIED — why it did not count against the coach (or did). "
                 "'CLEAR DECLINE' = real 5v5 chance, chose not to. Chance quality = share of the pull zone "
                 "he had cleanly available at 5v5 (>=70% = clear chance). Filter any column.")
    npl["A1"].font = Font(name="Arial", bold=True, size=11)
    npl.merge_cells("A1:I1")
    nhdr = ["Season", "Date", "Team", "Coach", "Gap-3 from", "Until", "Chance quality (5v5)", "Reason", "Blocked by"]
    for j, h in enumerate(nhdr, 1):
        c = npl.cell(row=2, column=j, value=h); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    REASON_FILL = {
        "CLEAR DECLINE": PatternFill("solid", fgColor="F8696B"),
        "partial chance": PatternFill("solid", fgColor="FFEB84"),
    }
    nr = 3
    for r in sorted(cwr, key=lambda x: (x["date"], x["coach"])):
        if r["pulled"]:
            continue
        blk = ", ".join(f"{k} {v}s" for k, v in r["blocked"].items()) or ""
        o = r["opened_R"]; c_ = r["closed_R"]
        vals = [r["season"][:4] + "-" + r["season"][6:], r["date"], r["team"], r["coach"],
                ("period start (carry-in)" if o >= 1200 else f"{o//60}:{o%60:02d}"),
                ("horn" if c_ == 0 else f"{c_//60}:{c_%60:02d}"),
                r["frac_ev"], r["reason"], blk]
        for j, v in enumerate(vals, 1):
            c = npl.cell(row=nr, column=j, value=v)
            c.font = ARIAL; c.border = THIN
            if j == 7: c.number_format = "0%"
            if j == 8 and r["reason"] in REASON_FILL: c.fill = REASON_FILL[r["reason"]]
        nr += 1
    lastn = nr - 1
    npl.auto_filter.ref = f"A2:I{lastn}"
    npl.freeze_panes = "A3"
    for col, w in zip("ABCDEFGHI", (9, 11, 7, 24, 13, 8, 13, 24, 34)):
        npl.column_dimensions[col].width = w

    # legend sheet
    lg = wb.create_sheet("HOW TO USE")
    notes = [
        "HOW TO USE (3 steps)",
        "1. COACH GUIDE tab: find the trailing team -> current coach (bold). Read his pull % on real chances, his typical pull time, and which LINES row to use.",
        "2. LINES tab: filter 'Situation' + 'Coach type' to what the guide told you, find the time remaining.",
        "3. TRUE line = fair price (break-even). Line @10% = worst price that still pays +10%% EV. Book offering better than the @10%% number = bet. Worse = pass.",
        "",
        "WHY PP PULLS ARE SHOWN SEPARATELY (NOT a betting trigger):",
        "You will never get a line once the trailing team is on a power play — books pull the market. So PP pulls can never be bet on.",
        "They matter for one reason: getting a coach's TRUE 5v5 willingness right. Four coaches (Tortorella, McLellan, Cronin, Huska)",
        "look like occasional pullers in raw stats, but every one of their pulls came on a PP. Counting those would trick you into",
        "pricing some 5v5 pull chance for them. The truth: at 5v5 they are 0-for-23 combined on clean chances — treat as no-pull teams.",
        "Same logic everywhere: the guide's pull % counts ONLY clean 5v5 chances, so e.g. Keefe is 38% (not the 53% raw) and",
        "Bednar's 100% is 6/6 pure 5v5. The POWER PLAY rows stay in the LINES tab for understanding the game, not for betting.",
        "",
        "Reading lines: -196 means bet only at -196 or better (e.g. -180, -150, +110). +150 means +150 or longer.",
        "Pull % caveat: 9/9 does NOT mean a guaranteed 100%% — small samples, read the Range column (9/9 -> can't rule out ~82%).",
        "Colors: green = likely, red = unlikely. Blue rows = trailing team on power play. Orange rows = net already empty.",
        "NO-GO for real money (Seb ruling). Leader market = Overs only. Provenance: MC grid n=30k seed 7, blind-validated 25-26.",
        "",
        "TIMING IS SOFT INFO: pulls happen at opportunities (offensive-zone faceoff, puck control), so an observed avg like 7:30",
        "off 9 pulls is NOT his true intent time. Team pricing shrinks each coach's timing ~40% toward league (4-season persistence 0.608)",
        "plus a small-sample discount. And a proven puller who is past his usual time with the net still in is still COMING —",
        "he likely just hasn't had the opportunity moment. Downgrade gradually, not instantly (the model conditions for this softly).",
        "Rebuild (cheap session): python3 hockeycore/io/build_lines_table.py  (needs data/derived/coach_profiles.json)",
    ]
    for i, t in enumerate(notes, 1):
        lg.cell(row=i, column=1, value=t).font = BOLD if i == 1 else ARIAL
    lg.column_dimensions["A"].width = 130

    wb.save(OUT)
    print("saved", OUT, f"({last-2} data rows)")


if __name__ == "__main__":
    main()
