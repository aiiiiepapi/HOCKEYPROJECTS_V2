"""Build 3mapgot_calculator_v2.xlsx (version 3: per-second 10%-EV lines).

Sheets:
  README  — provenance, rulings, warnings
  LOOKUP  — 6 yellow inputs (time, state, coach m, home, fav/dog, edge)
            -> bilinear interpolation (R mesh 20s x coach tier) -> per-market
            probability AND the American line that clears the edge
  PRICES  — dense MC grid, 37 R x 3 states x 6 tiers (n=30k, seed 7)
  COACHES — production multipliers + RISKY flag (<5 real chances, Seb ruling)
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
DER = ROOT / "data" / "derived"
OUT = ROOT / "live_model" / "3mapgot_calculator_v2.xlsx"

YELLOW = PatternFill("solid", fgColor="FFFF00")
BOLD = Font(name="Arial", bold=True)
ARIAL = Font(name="Arial")
BLUE = Font(name="Arial", color="0000FF")
RED = Font(name="Arial", color="FF0000", bold=True)

grid = json.load(open(DER / "pricing_grid_dense.json"))
manual = json.load(open(DER / "pricing_grid_manual.json"))
coaches = json.load(open(DER / "coach_table_production.json"))
chances = json.load(open(DER / "coach_chances.json"))
_lt = json.load(open(DER / "low_tail_coaches.json")) if (DER / "low_tail_coaches.json").exists() else []
low_tail = set(_lt["coaches"] if isinstance(_lt, dict) else _lt)
roi = json.load(open(DER / "roi_at_threshold.json"))["markets"]

wb = Workbook()

# ---------- README ----------
ws = wb.active
ws.title = "README"
lines = [
    "3MAPGOT CALCULATOR v2.3 — per-second 10%-EV lines (built from dense MC grid, n=30k, seed 7)",
    "Provenance: fits.json (both seasons, exposure-based), coach attenuation 0.355, context multipliers home 1.34 / fav 1.07 / dog 0.64.",
    "",
    "WHAT'S NEW (Seb rulings 2026-07-25):",
    "- Any time between 3:00 and 15:00 accepted (bilinear interpolation on a 20-second mesh).",
    "- LINE output per market: the American line at which the bet clears the edge in cell D8 (default +10% EV). Bet only at that number or better.",
    "- Coach policy: NO stake mandates. Coaches with <5 real chances are flagged RISKY in COACHES col G — your sizing call.",
    "- Status: NO-GO for real money (Seb). This is a research/paper tool until 99% confidence.",
    "",
    "VALIDATION (blind 25-26, betting EVERY spot at the exact 10%-EV line, game-cluster CIs):",
    f"- leader TT over: realized ROI {roi['leaderTT_over']['roi']:+.1%}, CI [{roi['leaderTT_over']['ci95_game_cluster'][0]:+.1%}, {roi['leaderTT_over']['ci95_game_cluster'][1]:+.1%}] — PROVEN >10%",
    f"- game total over: realized ROI {roi['gametotal_over']['roi']:+.1%}, CI [{roi['gametotal_over']['ci95_game_cluster'][0]:+.1%}, {roi['gametotal_over']['ci95_game_cluster'][1]:+.1%}] — positive, 10% not proven",
    f"- leader -3.5: realized ROI {roi['leader_-3.5']['roi']:+.1%}, CI [{roi['leader_-3.5']['ci95_game_cluster'][0]:+.1%}, {roi['leader_-3.5']['ci95_game_cluster'][1]:+.1%}] — CI includes 0: treat edges as optimistic",
    "",
    "STANDING WARNINGS:",
    "- Leader market: model runs ~5pts LOW (under investigation). Overs ONLY — never bet leader-market unders off this sheet.",
    "- DROPPED (Seb 2026-07-25): all 2+-more-goals markets. Not part of the product.",
    "- Low-tail coaches (Huska, Tortorella, McLellan): home/fav context multipliers do NOT apply — enter 'n'/'even' for them.",
    "- Timing: Roy-type late pullers shift the curve; production shifts applied in the fits, not user-adjustable here.",
]
for i, t in enumerate(lines, 1):
    ws.cell(row=i, column=1, value=t).font = BOLD if i == 1 else ARIAL
ws.column_dimensions["A"].width = 140

# ---------- PRICES (dense) ----------
ws = wb.create_sheet("PRICES")
hdr = ["key", "R", "time", "state", "tier", "P_total_ge1", "P_total_ge2",
       "P_leader_ge1", "P_margin_ge4", "P_leader_ge2"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=1, column=j, value=h).font = BOLD
r = 2
for row in sorted(grid, key=lambda x: (-x["R"], x["state"], x["tier"])):
    key = f'{row["R"]}|{row["state"]}|{row["tier"]:.2f}'
    vals = [key, row["R"], f'{row["R"] // 60}:{row["R"] % 60:02d}', row["state"], row["tier"],
            row["P_total_ge1"], row["P_total_ge2"], row["P_leader_ge1"],
            row["P_margin_ge4"], row["P_leader_ge2"]]
    for j, v in enumerate(vals, 1):
        ws.cell(row=r, column=j, value=v).font = ARIAL
    r += 1

# ---------- LOOKUP ----------
ws = wb.create_sheet("LOOKUP")
ws.column_dimensions["A"].width = 62
for c in "BCDEF":
    ws.column_dimensions[c].width = 13

def put(cell, val, font=ARIAL, fill=None, fmt=None):
    c = ws[cell]
    c.value = val
    c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt

put("A1", "LIVE LOOKUP — fill the 6 yellow cells", BOLD)
put("A3", "Time remaining (any mm:ss from 3:00 to 15:00)"); put("D3", "10:00", BLUE, YELLOW)
put("A4", "State (not_pulled_EV / not_pulled_PP / pulled)"); put("D4", "not_pulled_EV", BLUE, YELLOW)
put("A5", "Coach PRODUCTION m (COACHES col E; unknown = 1.00)"); put("D5", 1.0, BLUE, YELLOW)
put("A6", "Trailing team at HOME? (y/n — 'n' for low-tail coaches)"); put("D6", "n", BLUE, YELLOW)
put("A7", "Trailing team is... (fav / even / dog) by standings"); put("D7", "even", BLUE, YELLOW)
put("A8", "Edge required (0.10 = +10% EV)"); put("D8", 0.10, BLUE, YELLOW, "0%")

put("A9", "R secs / m_total:")
put("D9", '=MAX(180,MIN(900,VALUE(LEFT(D3,FIND(":",D3)-1))*60+VALUE(MID(D3,FIND(":",D3)+1,2))))')
put("E9", '=D5*IF(LOWER(D6)="y",1.34,1)*IF(LOWER(D7)="fav",1.07,IF(LOWER(D7)="dog",0.64,1))')
put("A10", "tier lo/hi/weight:")
put("D10", '=IF(E9<=0.8,0.55,IF(E9<=1,0.8,IF(E9<=1.3,1,IF(E9<=1.85,1.3,1.85))))')
put("E10", '=IF(E9<=0.8,0.8,IF(E9<=1,1,IF(E9<=1.3,1.3,IF(E9<=1.85,1.85,2.4))))')
put("F10", '=MAX(0,MIN(1,(E9-D10)/(E10-D10)))')
put("A11", "R lo/hi/weight:")
put("D11", '=MIN(880,FLOOR(D9,20))')
put("E11", '=D11+20')
put("F11", '=MAX(0,MIN(1,(D9-D11)/20))')
put("A12", "corner keys:")
put("B12", '=D11&"|"&D4&"|"&TEXT(D10,"0.00")')
put("C12", '=D11&"|"&D4&"|"&TEXT(E10,"0.00")')
put("D12", '=E11&"|"&D4&"|"&TEXT(D10,"0.00")')
put("E12", '=E11&"|"&D4&"|"&TEXT(E10,"0.00")')

put("A14", "MARKET", BOLD); put("D14", "PROBABILITY", BOLD); put("E14", "LINE for +edge EV", BOLD)
markets = [
    ("P(any more goal) — game total over", "F", 15),
    ("P(LEADER scores again) — team total over (model ~5pts LOW; Overs only)", "H", 16),
    ("P(leader wins by 4+) — the -3.5", "I", 17),
]
# helper corner block rows 22-25: cols B..E = corners (Rlo,tlo) (Rlo,thi) (Rhi,tlo) (Rhi,thi)
put("A21", "helper: corner probabilities (do not edit)", BOLD)
keycell = {"B": "$B$12", "C": "$C$12", "D": "$D$12", "E": "$E$12"}
for label, col, orow in markets:
    hrow = orow + 7  # 22..25
    put(f"A{hrow}", label.split(" — ")[0])
    for cc in "BCDE":
        put(f"{cc}{hrow}", f'=INDEX(PRICES!{col}:{col},MATCH({keycell[cc]},PRICES!A:A,0))', fmt="0.0000")
    put(f"A{orow}", label)
    put(f"D{orow}",
        f'=(1-$F$11)*((1-$F$10)*B{hrow}+$F$10*C{hrow})+$F$11*((1-$F$10)*D{hrow}+$F$10*E{hrow})',
        fmt="0.0%")
    put(f"E{orow}",
        f'=IF(D{orow}<=0,"n/a",IF((1+$D$8-D{orow})/D{orow}>=1,'
        f'"+"&TEXT(ROUND((1+$D$8-D{orow})/D{orow}*100,0),"0"),'
        f'"-"&TEXT(ROUND(100/((1+$D$8-D{orow})/D{orow}),0),"0")))', BOLD)

put("A19", "Warnings:")
put("D19", '=IF(AND(D4="pulled",D9>420),"EXTRAPOLATED — no real pulls this early","ok")', RED)


# ---------- MANUAL_PRICES (hidden data) ----------
ws = wb.create_sheet("MANUAL_PRICES")
mh = ["key", "R", "pullpct", "shift", "P_total_ge1", "P_leader_ge1", "P_margin_ge4"]
for j, h in enumerate(mh, 1):
    ws.cell(row=1, column=j, value=h).font = BOLD
r = 2
for row in sorted(manual, key=lambda x: (-x["R"], x["pullpct"], x["shift"])):
    key = f'{row["R"]}|{row["pullpct"]:.3f}|{row["shift"]}'
    for j, v in enumerate([key, row["R"], row["pullpct"], row["shift"],
                           row["P_total_ge1"], row["P_leader_ge1"], row["P_margin_ge4"]], 1):
        ws.cell(row=r, column=j, value=v).font = ARIAL
    r += 1
ws.sheet_state = "hidden"

# ---------- MANUAL (Seb's override inputs) ----------
ws = wb.create_sheet("MANUAL")
ws.column_dimensions["A"].width = 60
for c in "BCDEF":
    ws.column_dimensions[c].width = 13

def putm(cell, val, font=ARIAL, fill=None, fmt=None):
    c = ws[cell]; c.value = val; c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt

putm("A1", "MANUAL COACH OVERRIDE — your own pull read, priced by the model", BOLD)
putm("A3", "Time remaining (any mm:ss from 3:00 to 15:00)"); putm("D3", "6:00", BLUE, YELLOW)
putm("A4", "YOUR expected pull % on this real chance (0.10-0.99)"); putm("D4", 0.695, BLUE, YELLOW, "0%")
putm("A5", "His typical pull time left (league = 4:19)"); putm("D5", "4:19", BLUE, YELLOW)
putm("A6", "Edge required"); putm("D6", 0.10, BLUE, YELLOW, "0%")
putm("A8", "R / shift secs (snapped to grid):")
putm("D8", '=MAX(180,MIN(900,VALUE(LEFT(D3,FIND(":",D3)-1))*60+VALUE(MID(D3,FIND(":",D3)+1,2))))')
putm("E8", '=MAX(-60,MIN(180,VALUE(LEFT(D5,FIND(":",D5)-1))*60+VALUE(MID(D5,FIND(":",D5)+1,2))-259))')
putm("F8", '=INDEX({-60;0;60;120;180},MATCH(MIN(ABS({-60;0;60;120;180}-E8)),ABS({-60;0;60;120;180}-E8),0))')
putm("A9", "R lo/hi/w:")
putm("D9", '=MIN(880,FLOOR(D8,20))'); putm("E9", '=D9+20'); putm("F9", '=MAX(0,MIN(1,(D8-D9)/20))')
putm("A10", "pull%% lo/hi/w:")
PG = "{0.1;0.3;0.5;0.695;0.85;0.95;0.99}"
putm("D10", f'=MAX(0.1,INDEX({PG},MATCH(MIN(MAX(D4,0.1),0.99),{PG},1)))')
putm("E10", f'=MIN(0.99,INDEX({PG},MIN(MATCH(D10,{PG},0)+1,7)))')
putm("F10", '=IF(E10=D10,0,MAX(0,MIN(1,(MIN(MAX(D4,0.1),0.99)-D10)/(E10-D10))))')
putm("A11", "corner keys:")
putm("B11", '=D9&"|"&TEXT(D10,"0.000")&"|"&F8')
putm("C11", '=D9&"|"&TEXT(E10,"0.000")&"|"&F8')
putm("D11", '=E9&"|"&TEXT(D10,"0.000")&"|"&F8')
putm("E11", '=E9&"|"&TEXT(E10,"0.000")&"|"&F8')
putm("A13", "MARKET", BOLD); putm("D13", "PROBABILITY", BOLD); putm("E13", "LINE for +edge EV", BOLD)
mmk = [("Game total OVER (any more goal)", "E", 14),
       ("Leader TT OVER (Overs only; model ~5pts low)", "F", 15),
       ("Leader -3.5 (wins by 4+)", "G", 16)]
for label, col, orow in mmk:
    hrow = orow + 7
    putm(f"A{hrow}", label.split(" (")[0])
    for cc, keyc in (("B","$B$11"),("C","$C$11"),("D","$D$11"),("E","$E$11")):
        putm(f"{cc}{hrow}", f'=INDEX(MANUAL_PRICES!{col}:{col},MATCH({keyc},MANUAL_PRICES!A:A,0))', fmt="0.0000")
    putm(f"A{orow}", label)
    putm(f"D{orow}", f'=(1-$F$9)*((1-$F$10)*B{hrow}+$F$10*C{hrow})+$F$9*((1-$F$10)*D{hrow}+$F$10*E{hrow})', fmt="0.0%")
    putm(f"E{orow}", f'=IF(D{orow}<=0,"n/a",IF((1+$D$6-D{orow})/D{orow}>=1,'
                     f'"+"&TEXT(ROUND((1+$D$6-D{orow})/D{orow}*100,0),"0"),'
                     f'"-"&TEXT(ROUND(100/((1+$D$6-D{orow})/D{orow}),0),"0")))', BOLD)
putm("A18", "Notes: timing snapped to nearest grid step (max 30s, ~1pt effect). 100%% pullers: use 0.99.", ARIAL)
putm("A19", "5v5 only — PP/pulled states not bettable (Seb rulings).", ARIAL)

# ---------- COACHES ----------
ws = wb.create_sheet("COACHES")
hdr = ["coach", "pulls", "expected", "m_EB", "m_PRODUCTION (use this)", "real chances", "flag"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=1, column=j, value=h).font = BOLD
ws.column_dimensions["A"].width = 24
ws.column_dimensions["E"].width = 22
r = 2
for c in sorted(coaches["coaches"], key=lambda x: -x["m_prod"]):
    n_ch = chances.get(c["coach"], 0)
    flags = []
    if n_ch < 5:
        flags.append("RISKY (<5 chances)")
    if c["coach"] in low_tail:
        flags.append("LOW-TAIL: context mults OFF")
    vals = [c["coach"], c["pulls"], c["expected"], c["m_eb"], c["m_prod"], n_ch, "; ".join(flags)]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = RED if (j == 7 and flags) else ARIAL
    r += 1
ws.cell(row=r + 1, column=1,
        value="Seb ruling 2026-07-25: no stake mandates — RISKY flag is informational. "
              "Flag resets when a coach changes team (<5 chances with current team).").font = ARIAL

wb.save(OUT)
print("saved", OUT)
