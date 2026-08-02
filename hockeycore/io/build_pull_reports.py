"""
build_pull_reports.py — the three per-league pull reports (xlsx).

Layout (Seb spec 2026-08-02): Team | Coach | Expected pull % |
Career (pulls/chances) | Pulls 2025-26 | No pulls 2025-26.
Active coaches only (last_seen >= 2025-06-01). AHL sheet carries the
no-priced-markets footer (ruling 24).

Usage: python3 hockeycore/io/build_pull_reports.py [outdir]
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[2]
DER = ROOT / "data" / "derived"
ACTIVE = "2025-06-01"
SEASON = "2025-07-01"

CFG = {
    "nhl": ("NHL", "coach_profiles.json", None),
    "ahl": ("AHL", "ahl_coach_profiles.json", "ahl_clean_window.json"),
    "liiga": ("Liiga", "liiga_coach_profiles.json", "liiga_clean_window.json"),
}


def season_counts(league):
    """(coach -> (pulls, nopulls)) on 2025-26 clean chances."""
    out = {}
    if league == "nhl":
        rows = json.load(open(DER / "clean_window_instances.json"))
        for r in rows:
            if r["date"] < SEASON:
                continue
            took = r["pulled"] and r["pull_type"] == "ev"
            dec = (not r["pulled"]) and r["frac_ev"] >= 0.7
            if took or dec:
                p, n = out.get(r["coach"], (0, 0))
                out[r["coach"]] = (p + int(took), n + int(dec))
    else:
        rows = json.load(open(DER / CFG[league][2]))["rows"]
        for r in rows:
            if r["date"] >= SEASON and r["clear"]:
                p, n = out.get(r["coach"], (0, 0))
                out[r["coach"]] = (p + int(r["taken"]), n + int(not r["taken"]))
    return out


def build(league, outdir):
    label, prof_f, _ = CFG[league]
    raw = json.load(open(DER / prof_f))
    profs = raw["profiles"] if isinstance(raw, dict) else raw
    profs = [p for p in profs if p["last_seen"] >= ACTIVE]
    profs.sort(key=lambda p: -p["expected_pull_pct"])
    sc = season_counts(league)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{label} pulls"
    ws.append([f"{label} — expected pull % (trailing by 3, 3rd period). "
               "% is career-wide, ~12-month half-weight on recency; "
               "hot coaches (75%+ on 2+ chances) near-unanchored (ruling 33)."])
    ws["A1"].font = Font(italic=True, size=9)
    ws.append(["Team", "Coach", "Expected pull %", "Career (pulls/chances)",
               "Pulls 2025-26", "No pulls 2025-26"])
    for c in ws[2]:
        c.font = Font(bold=True)
    for p in profs:
        yp, yn = sc.get(p["coach"], (0, 0))
        ws.append([p["team"], p["coach"], round(p["expected_pull_pct"], 3),
                   f'{p["clear_taken"]}/{p["clear_chances"]}', yp, yn])
        ws.cell(ws.max_row, 3).number_format = "0%"
    for col, w in zip("ABCDEF", (8, 24, 15, 20, 13, 15)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(horizontal="center" if c.column > 2 else "left")
    if league == "ahl":
        ws.append([])
        ws.append(["AHL has NO priced markets (ruling 24) — coach intel only."])
        ws.cell(ws.max_row, 1).font = Font(italic=True, size=9)
    out = Path(outdir) / f"pull_report_{league}_2025-26.xlsx"
    wb.save(out)
    print(f"wrote {out} ({len(profs)} coaches)")


if __name__ == "__main__":
    outdir = sys.argv[1] if len(sys.argv) > 1 else "/home/claude/work"
    for lg in CFG:
        build(lg, outdir)
